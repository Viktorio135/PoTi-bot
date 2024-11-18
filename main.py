import json
import logging
import threading
import os
import time
import random
import openpyxl

from dotenv import load_dotenv
from apscheduler.schedulers.asyncio import AsyncIOScheduler

from aiogram import Bot, Dispatcher, executor, types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Text
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.utils.deep_linking import get_start_link, decode_payload
from aiogram.types import ContentType

from dataBase.db_commands import (
    create_user, has_register, 
    get_user_by_id, delete_profile, 
    get_list_of_profiles, update_active_to_true, update_active_to_false,
    change_description_by_id,
    get_list_of_admins, block_user_db,
    unblock_user_db, get_statistic_user_db, get_list_of_users_for_spam_db,
    add_admin_db, delete_admin_db, update_screenshots, 
    get_ref_stat, delete_user
    
)
from keyboards import ( 
    end_registration_kb, menu_kb,
    my_profile_kb, search_kb,
    show_like_kb, like_kb, report_kb,
    change_profile_description_cancel,
    change_profile_photo_cancel, check_sub, second_screenshot, change_profile_photo2_cancel, reminder_kb,
    select_role, skip_description
)

from states.user_states import (
    Register_new_user, 
    Change_photo, ReportUserOther, Change_description
)

from states.admin_states import (
    AdminBlockUser, AdminUnblockUser, AdminGetUserById,
    AdminSpamHasPhoto, AdminSpamOnlyText, AdminSpamWithPhoto,
    AdminGetUserByPhoto, AdminSendMessageFromUserById, AdminAddAdmin,
    AdminDeleteAdmin, AdminDeleteUser, AdminGetRefStat
)

from dataBase.dump import dump_dict_of_profiles, backup_bd
from utils.search_photo import compare_images
from dataBase.models import start_db
from utils.scheduler import start_schedule




load_dotenv()
token = os.getenv('TOKEN')
main_admin = os.getenv('MAIN_ADMIN')

bot = Bot(token)
dp = Dispatcher(bot, storage=MemoryStorage())

cached_data = {}
dict_of_profiles = {}

love_activity = [0]



@dp.message_handler(commands='start')
async def cmd_start(msg: types.Message):

    args = msg.get_args()
    if args != '':
        cached_data[msg.from_user.id] = args
    else:
        cached_data[msg.from_user.id] = ''

    user_channel_status = await bot.get_chat_member(
            # chat_id='-1002171208182',
            chat_id='-1002239230673', 
            user_id=msg.from_user.id
            )

    if user_channel_status.status != 'left':
        await bot.send_message(
                msg.from_user.id, 
                'Привет, ты попал в ПоТи MLBB\n\nДля начала использования подпишитесь на канал https://t.me/PoTi_MLBB\n\nМы поможем тебе найти кого-нибудь в Mobile legends.',
                reply_markup=types.ReplyKeyboardRemove()
                )
        if await has_register(str(msg.from_user.id)):
            await menu(msg)
        elif not(await has_register(str(msg.from_user.id))):
            await register_or_update_user(msg)
    else:
        await bot.send_message(
            msg.from_user.id,
            'Привет, ты попал в ПоТи MLBB\n\nДля начала использования подпишитесь на канал https://t.me/PoTi_MLBB\n\nМы поможем тебе найти кого-нибудь в Mobile legends.\n\nДля начала использования подпишитесь на канал!',
            reply_markup=check_sub()
        )


@dp.callback_query_handler(lambda c: c.data == 'check_sub')
async def check_sub_start(callback_query: types.CallbackQuery):
    user_channel_status = await bot.get_chat_member(
            # chat_id='-1002171208182',
            chat_id='-1002239230673', 
            user_id=callback_query.from_user.id
            )
    if user_channel_status.status != 'left':
        await bot.send_message(callback_query.from_user.id, 'Спасибо за подписку!')
        if await has_register(str(callback_query.from_user.id)):
            await menu(callback_query)
        elif not(await has_register(str(callback_query.from_user.id))):
            await register_or_update_user(callback_query)
    else:
        await bot.send_message(
            callback_query.from_user.id,
            'Для начала использования подпишитесь на канал!',
            reply_markup=check_sub()
        )


############################## Регистрация ####################################

async def register_or_update_user(msg: types.Message, is_new=False):
    user_id = str(msg.from_user.id)
    if not is_new:
        await bot.send_message(
                msg.from_user.id, 
                'Для начала, давай создадим анкету!Напишите мне свой игровой никнейм, которое будут видеть все.',
                reply_markup=types.ReplyKeyboardRemove()
                )
        await Register_new_user.name.set()
    else:
        await bot.send_message(
                msg.from_user.id, 
                'Давай создадим анкету заново.\nНапишите мне свой игровой никнейм, которое будут видеть все.',
                reply_markup=types.ReplyKeyboardRemove()
                )
        await Register_new_user.name.set()
         

@dp.message_handler(state=Register_new_user.name)
async def register_name(msg: types.Message, state: FSMContext):
    if msg.text == '/start':
        await state.finish()
        await cmd_start(msg)
    else:
        async with state.proxy() as data:
            data['name'] = msg.text

        await bot.send_message(
            msg.from_user.id, 
            'Напиши свой игровой ID.', 
            )
        await Register_new_user.next()



@dp.message_handler(state=Register_new_user.game_id)
async def register_sex(msg: types.Message, state: FSMContext):
    if msg.text == '/start':
        await state.finish()
        await cmd_start(msg)
    else:
        async with state.proxy() as data:
            data['game_id'] = msg.text

        await bot.send_message(
            msg.from_user.id, 
            'Теперь выбери роль (до 3)',
            reply_markup=select_role()
            # 'Расскажи о своем игровом опыте и кого хочешь найти. Это поможет лучше подобрать тебе тимейтов.',
            )
        await Register_new_user.next()


@dp.message_handler(state=Register_new_user.role1)
async def register_role_1(msg: types.Message, state: FSMContext):
    roles = ['Золото 🥇', 'Опыт 💪🏼', 'Мид 🧙🏻‍♂️', 'Лес 🌳','Роум 🔰','Все']
    if msg.text == '/start':
        await state.finish()
        await cmd_start(msg)
    else:
        if msg.text in roles:
            if msg.text == 'Все':
                async with state.proxy() as data:
                    data['role1'] = 'Все'
                    data['role2'] = ''
                    data['role3'] = ''
                await bot.send_message(
                    msg.from_user.id,
                    'Расскажи о своем игровом опыте и кого хочешь найти. Это поможет лучше подобрать тебе тимейтов.',
                    reply_markup=skip_description()
                )
                await Register_new_user.description.set()
            else:
                async with state.proxy() as data:
                    data['role1'] = msg.text
                await bot.send_message(
                    msg.from_user.id,
                    'Роль добавлена 1 из 3, Добавить еще одно ?',
                    reply_markup=select_role(second=True)
                )
                await Register_new_user.next()
        else:
            await bot.send_message(
                msg.from_user.id,
                'Нет такого варианта ответа, попробуйте еще раз!'
            )
            return


@dp.message_handler(state=Register_new_user.role2)
async def register_role_2(msg: types.Message, state: FSMContext):
    roles = ['Золото 🥇', 'Опыт 💪🏼', 'Мид 🧙🏻‍♂️', 'Лес 🌳','Роум 🔰', 'Все', 'Это все, сохранить роли']
    if msg.text == '/start':
        await state.finish()
        await cmd_start(msg)
    else:
        async with state.proxy() as data:
            roles.remove(data['role1'])
        if msg.text in roles:
            if msg.text != 'Это все, сохранить роли':
                async with state.proxy() as data:
                    data['role2'] = msg.text
                await bot.send_message(
                    msg.from_user.id,
                    'Роль добавлена 2 из 3, выберите еще одну роль, либо нажмите "Это все, сохранить роли".',
                    reply_markup=select_role(second=True)
                )
                await Register_new_user.next()
            else:
                async with state.proxy() as data:
                    data['role2'] = ''
                    data['role3'] = ''
                    await bot.send_message(
                        msg.from_user.id,
                        'Расскажи о своем игровом опыте и кого хочешь найти. Это поможет лучше подобрать тебе тимейтов.',
                        reply_markup=skip_description()
                    )
                    await Register_new_user.description.set()
                
        else:
            await bot.send_message(
                msg.from_user.id,
                'Нет такого варианта ответа, либо данная роль уже выбрана, попробуйте еще раз!'
            )
            return

@dp.message_handler(state=Register_new_user.role3)
async def register_role_3(msg: types.Message, state: FSMContext):
    roles = ['Золото 🥇', 'Опыт 💪🏼', 'Мид 🧙🏻‍♂️', 'Лес 🌳','Роум 🔰', 'Все', 'Это все, сохранить роли']
    if msg.text == '/start':
        await state.finish()
        await cmd_start(msg)
    else:
        async with state.proxy() as data:
            roles.remove(data['role1'])
            roles.remove(data['role2'])
        if msg.text in roles:
            if msg.text != 'Это все, сохранить роли':
                async with state.proxy() as data:
                    data['role3'] = msg.text
                    await bot.send_message(
                        msg.from_user.id,
                        'Расскажи о своем игровом опыте и кого хочешь найти. Это поможет лучше подобрать тебе тимейтов.',
                        reply_markup=skip_description()
                    )
                    await Register_new_user.next()
            else:
                async with state.proxy() as data:
                    data['role3'] = ''
                    await bot.send_message(
                        msg.from_user.id,
                        'Расскажи о своем игровом опыте и кого хочешь найти. Это поможет лучше подобрать тебе тимейтов.',
                        reply_markup=skip_description()
                    )
                    await Register_new_user.next()
                
        else:
            await bot.send_message(
                msg.from_user.id,
                'Нет такого варианта ответа, либо данная роль уже выбрана, попробуйте еще раз!'
            )
            return

@dp.message_handler(state=Register_new_user.description)
async def register_search(msg: types.Message, state: FSMContext):
    if msg.text == '/start':
        await state.finish()
        await cmd_start(msg)
    else:
        if msg.text != 'Пропустить':
            if len(msg.text) < 800:
                if ('https://' not in msg.text) and ('.com' not in msg.text) and ('.ru' not in msg.text) and ('t.me' not in msg.text):
                    async with state.proxy() as data:
                        data['description'] = msg.text

                    await bot.send_message(
                        msg.from_user.id, 
                        'Теперь пришли скриншот профиля, его будут видеть другие пользователи. Обязательно загружайте по одной фотографии.🤏',
                        reply_markup=types.ReplyKeyboardRemove()
                        )
                    await Register_new_user.next()
                else:
                    await bot.send_message(
                        msg.from_user.id,
                        'Этот текст нарушает правила, введите другой'
                    )
                    return
            else:
                await bot.send_message(
                    msg.from_user.id,
                    'Описание слишком длинное, попробуйте ввести другое (до 800 символов)',
                )
                return
        else:
            async with state.proxy() as data:
                data["description"] = ''
            await bot.send_message(
                    msg.from_user.id, 
                    'Теперь пришли скриншот профиля, его будут видеть другие пользователи. Обязательно загружайте по одной фотографии.🤏',
                    reply_markup=types.ReplyKeyboardRemove()
                    )
            await Register_new_user.next()


@dp.message_handler(content_types=[ContentType.PHOTO, ContentType.TEXT], state=Register_new_user.screenshot1)
async def register_photo1(msg: types.Message, state: FSMContext):
    user_id = str(msg.from_user.id)
    if msg.text == '/start':
        await state.finish()
        await cmd_start(msg)
    else:
        try:
            file_name = f'{msg.from_user.id}_1.jpg'
            path = f'static/users_photo/{file_name}'
            await msg.photo[-1].download(path)

            async with state.proxy() as data:
                    data['screenshot1'] = path
                    
            await bot.send_message(
                msg.from_user.id, 
                'Фото добавлено 1 из 2, Есть еще одно ?',
                reply_markup=second_screenshot()
                )
            await Register_new_user.next()

        except Exception as e:   
            logging.error(f'Ошибка в регистрации при сохранении фото у пользователя {user_id}', exc_info=True) 
            await bot.send_message(
                msg.from_user.id, 
                'Возникла ошибка. Попробуйте еще раз'
                )
            await state.finish()
            await register_or_update_user(msg)
        

@dp.callback_query_handler(lambda c: c.data == 'second_screenshot', state=Register_new_user.screenshot2)
async def register_second_is_empty(callback_query: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        data["screenshot2"] = ''
    await state.finish()
    await end_registration(callback_query, data)
    


@dp.message_handler(content_types=[ContentType.PHOTO, ContentType.TEXT], state=Register_new_user.screenshot2)
async def register_photo2(msg: types.Message, state: FSMContext):
    user_id = str(msg.from_user.id)
    if msg.text == '/start':
        await state.finish()
        await cmd_start(msg)
    else:
        try:
        
            file_name = f'{msg.from_user.id}_2.jpg'
            path = f'static/users_photo/{file_name}'
            await msg.photo[-1].download(path)

            async with state.proxy() as data:
                    data['screenshot2'] = path
                    
            await state.finish()

            await end_registration(msg, data)
            
        except Exception as e:   
            logging.error(f'Ошибка в регистрации при сохранении фото у пользователя {user_id}', exc_info=True) 
            await bot.send_message(
                msg.from_user.id, 
                'Возникла ошибка. Попробуйте еще раз'
                )
            await state.finish()
            await register_or_update_user(msg)


async def end_registration(msg: types.Message, data):
    try:
        datas = {
            "name": data["name"],
            "game_id": data["game_id"],
            "description": data["description"], 
            "screenshot1": data["screenshot1"], 
            "screenshot2": data["screenshot2"], 
            "role1": data["role1"],
            "role2": data["role2"],
            "role3": data["role3"],
            "promocode": cached_data.get(msg.from_user.id),
            }
        if msg.from_user.id in cached_data:
            del cached_data[msg.from_user.id]
        file_id = hash(json.dumps(datas, sort_keys=True))
        cached_data[file_id] = datas


        await bot.send_message(msg.from_user.id, 
                f'Ты успешно завершил, твой профиль будет выглядеть вот так:'
            )
        if (msg.from_user.username == 'None') or (msg.from_user.username is None):
            await bot.send_message(
                msg.from_user.id,
    '''
    🚨🚨🚨🚨🚨🚨🚨🚨

    У вас в профиле телеграма не настроено имя пользователя (то что начинается с @). В связи с этим на вашу анкету не смогут отвечать другие пользователи. Для исправления этой проблемы необходимо установить имя пользователя в настройках телеграма и заполнить анкету в нашем боте заново!

    🚨🚨🚨🚨🚨🚨🚨🚨
    '''
            )
        
        media = types.MediaGroup()
        roles = '\nЛюбимые роли:\n'
        if data["role1"] != '':
            roles += f'{data["role1"]}\n'
        if data["role2"] != '':
            roles += f'{data["role2"]}\n'
        if data["role3"] != '':
            roles += f'{data["role3"]}\n'
        

        if (data["screenshot1"] != '') and (data["screenshot2"] != ''):
            media.attach_photo(types.InputFile(
                f'static/users_photo/{msg.from_user.id}_1.jpg'),
                f'{datas["name"]}\n{roles}\n{datas["description"]}')
            media.attach_photo(types.InputFile(f'static/users_photo/{msg.from_user.id}_2.jpg'))
            await bot.send_media_group(
                msg.from_user.id, 
                media=media,
            )

            
        elif (data["screenshot1"] != '') and (data["screenshot2"] == ''):
            media.attach_photo(types.InputFile(
                f'static/users_photo/{msg.from_user.id}_1.jpg'),
                f'{datas["name"]}\n{roles}\n{datas["description"]}')
            await bot.send_media_group(
                msg.from_user.id, 
                media=media,
            )

        else:
            await bot.send_message(
                msg.from_user.id,
                f'{datas["name"]}\n{roles}\n{datas["description"]}'
            )    

        await bot.send_message(
            msg.from_user.id,
            'Все верно?',
            reply_markup=end_registration_kb(file_id)
            )
    except:
        await bot.send_message(
            msg.from_user.id,
            'Что-то пошло не так, попробуйте еще раз...'
        )
        await register_or_update_user(msg)


    

@dp.callback_query_handler(lambda c: 'end_registration' in c.data)
async def save_user_to_bd(callback_query: types.CallbackQuery):
    try:
        file_id = int(callback_query.data.split(':')[1])
        data = cached_data.get(file_id)        
        del cached_data[file_id]

        user_name = callback_query.from_user.username
        id_user = callback_query.from_user.id
        if await has_register(id_user):
            await delete_profile(id_user)
        save = await create_user(
            name=data["name"], 
            game_id=data["game_id"], 
            user_id=id_user,
            screenshot1=data["screenshot1"], 
            screenshot2=data["screenshot2"],
            description=data["description"],
            role1=data["role1"],
            role2=data["role2"],
            role3=data["role3"],
            promocode=data["promocode"],
            username=user_name
        )


        if save:
            await bot.send_message(
                callback_query.from_user.id, 
                'Отлично! Надеюсь вы хорошо проведете время'
            )
            await bot.send_message(
                callback_query.from_user.id,
'''
Совет от ПоТи 
 Как не стать жертвой мошенников?
 Будь осторожнее когда после знакомства:

- тебя просят прислать личные данные, фотографии интимного характера. Их могут использовать против тебя: шантажировать и вымогать деньги.

- тебе прислали ссылку, после перехода по которой необходимо ввести личные данные / данные авторизации (логин/пароль). 

- тебя просят сделать покупку, например билеты в кино/театр, купить аккаунт игре. 

- тебя просят одолжить денег.

- тебе предлагают выгодную сделку, платные, инвестиционные и другие услуги.
'''

            )
            await bot.send_message(
                callback_query.from_user.id,
                'Немного подождите, идет подбор тимейтов...'
            )
            await bot.send_message(
                callback_query.from_user.id,
                '✨🔍',
                reply_markup=search_kb() 
            )
            await search_love_reg(callback_query)
            
        else:
            await bot.send_message(
                callback_query.from_user.id, 
                'Возникла ошибка. Попробуйте еще раз'
            )
            await register_or_update_user(callback_query)


    except Exception as e:
            logging.error(f'Ошибка в окончании регистрации у пользователя {str(callback_query.from_user.id)}', exc_info=True)
            await bot.send_message(
                 callback_query.from_user.id,
                 'Возникла ошибка. Попробуйте еще раз'
            )
            await register_or_update_user(callback_query)


@dp.callback_query_handler(lambda c: c.data == 'repeat_registration')
async def repeat_reg(callback_query: types.CallbackQuery):
    await register_or_update_user(callback_query, is_new=True)

##########################################################################################

#################################### Главное меню #####################################

@dp.message_handler(Text('💤'))
@dp.message_handler(commands='menu')
async def menu(msg: types.Message):
    try:
        user = await get_user_by_id(str(msg.from_user.id))
        
        if user != 'Error' and user != 'User not found':
            await bot.send_message(
                msg.from_user.id,
                '1. Смотреть анкеты.\n2. Моя анкета.\n3. Я больше не хочу никого искать.\n***\n4. Пригласи друзей - получи больше лайков 😎.',
                reply_markup=menu_kb()
            )

        else:
            await bot.send_message(
                msg.from_user.id,
                f'Возникла ошибка. Попробуйте еще раз',
            )
             
    except Exception as e:
        logging.error(f'Ошибка в главном меню у пользователя {str(msg.from_user.id)}', exc_info=True)
        await bot.send_message(
                msg.from_user.id,
                'Возникла ошибка. Попробуйте еще раз',
            )
        
@dp.message_handler(Text('𝟏'))
@dp.message_handler(Text('1'))
async def main_menu_1(msg: types.Message):
    await bot.send_message(
            msg.from_user.id,
            '✨🔍',
            reply_markup=search_kb() 
        )
    await search_love_reg(msg)



############# Анкета #############


@dp.message_handler(Text('2'))
async def my_profile(msg: types.Message):
    try:
        user_id = str(msg.from_user.id)
        is_blocked = await get_user_by_id(user_id)
        if not is_blocked["is_blocked"]:
            await bot.send_message(
                msg.from_user.id,
                'Ваша анкета выглядит вот так:'
            )
            user = await get_user_by_id(user_id)

            if user != 'User not found':
                media = types.MediaGroup()
                roles = '\nЛюбимые роли:\n'
                if user["role1"] != '':
                    roles += f'{user["role1"]}\n'
                if user["role2"] != '':
                    roles += f'{user["role2"]}\n'
                if user["role3"] != '':
                    roles += f'{user["role3"]}\n'

                if (user["screenshot1"] != '') and (user["screenshot2"] != ''):
                    media.attach_photo(types.InputFile(
                        f'static/users_photo/{msg.from_user.id}_1.jpg'),
                        f'{user["name"]}\n{roles}\n{user["description"]}'
                        )
                    media.attach_photo(types.InputFile(f'static/users_photo/{msg.from_user.id}_2.jpg'))
                    await bot.send_media_group(
                        msg.from_user.id, 
                        media=media,
                    )
                    

                    
                elif (user["screenshot1"] != '') and (user["screenshot2"] == ''):
                    media.attach_photo(types.InputFile(
                        f'static/users_photo/{msg.from_user.id}_1.jpg'),
                        f'{user["name"]}\n{roles}\n{user["description"]}')
                    await bot.send_media_group(
                        msg.from_user.id, 
                        media=media,
                    )

                else:
                    await bot.send_message(
                        msg.from_user.id,
                        f'{user["name"]}\n{roles}\n{user["description"]}'
                    )  
                await bot.send_message(
                    msg.from_user.id,
                    '1. Смотреть анкеты.\n2. Заполнить анкету заново.\n3. Изменить фото.\n4. Изменить текст анкеты.',
                    reply_markup=my_profile_kb()
                )  
            else:
                await delete_user(str(msg.from_user.id))
                if str(msg.from_user.id) in dict_of_profiles:
                    del dict_of_profiles[str(msg.from_user.id)]
                    for user in dict_of_profiles:
                        if str(msg.from_user.id) in dict_of_profiles[user]["profiles_list"]:
                            dict_of_profiles[user]["profiles_list"].remove(str(msg.from_user.id))
                await bot.send_message(
                    msg.from_user.id,
                    'Что-то пошло не так, введите /start и зарегистрируйтесь заново'
                )

        elif is_blocked["is_blocked"]:
            await bot.send_message(
                msg.from_user.id,
                'Ваша анкета была заблокирована'
            )
    except:
        await delete_user(str(msg.from_user.id))
        if str(msg.from_user.id) in dict_of_profiles:
            del dict_of_profiles[str(msg.from_user.id)]
            for user in dict_of_profiles:
                if str(msg.from_user.id) in dict_of_profiles[user]["profiles_list"]:
                    dict_of_profiles[user]["profiles_list"].remove(str(msg.from_user.id))
        await bot.send_message(
            msg.from_user.id,
            'Что-то пошло не так, введите /start и зарегистрируйтесь заново'
        )


@dp.message_handler(Text('𝟐'))
async def repeat_profile(msg: types.Message):
    user_id = str(msg.from_user.id)
    if user_id in dict_of_profiles:
        del dict_of_profiles[user_id]
    for users in dict_of_profiles:
        if user_id in dict_of_profiles[users]["profiles_list"]:
            dict_of_profiles[users]["profiles_list"].remove(user_id)
    await register_or_update_user(msg, is_new=True)



@dp.message_handler(Text('𝟑'))
async def change_photo(msg: types.Message):
    await bot.send_message(
        msg.from_user.id,
        'Пришли скриншот профиля, его будут видеть другие пользователи. ',
        reply_markup=change_profile_photo_cancel()
    )

    await Change_photo.screenshot1.set()

@dp.message_handler(content_types=['photo', 'text'], state=Change_photo.screenshot1)
async def change_photo_step1(msg: types.Message, state: FSMContext):
    if msg.content_type == 'text' and msg.text == '❌ Отмена':
        await state.finish()
        await my_profile(msg)
    if msg.content_type == 'text' and msg.text != '❌ Отмена':
        await bot.send_message(
            msg.from_user.id,
            'Нет такой комманды, отменить действие или отправьте фотографию!'
        )
        return
    else:
        file_name = f'{msg.from_user.id}_1.jpg'

        if os.path.exists(f'static/users_photo/{file_name}'):
            os.remove(f'static/users_photo/{file_name}')
        path = f'static/users_photo/{file_name}'
        await msg.photo[-1].download(path)

        async with state.proxy() as data:
                data['screenshot1'] = path
                
        await bot.send_message(
             msg.from_user.id, 
             'Фото добавлено 1 из 2, Есть еще одно ?',
             reply_markup=change_profile_photo2_cancel()
             )
        await Change_photo.next()

@dp.message_handler(content_types=['photo', 'text'], state=Change_photo.screenshot2)
async def change_photo_step2(msg: types.Message, state: FSMContext):
    if msg.content_type == 'text' and msg.text == 'Это все, сохранить фото':
        async with state.proxy() as data:
            data["screenshot2"] = ''
        if await update_screenshots(str(msg.from_user.id), screenshot1=data["screenshot1"], screenshot2=data["screenshot2"]):
            await state.finish()
            await bot.send_message(
                msg.from_user.id,
                'Успешно обновлено!'
            )
            await my_profile(msg)
        else:
            await bot.send_message(
                msg.from_user.id,
                'Что-то пошло не так, уже чиним...'
            )
            await state.finish()
            await my_profile(msg)
    if msg.content_type == 'text' and msg.text != 'Это все, сохранить фото':
        await bot.send_message(
            msg.from_user.id,
            'Нет такой команды, отправьте фотографию, либо сохраните первую!'
        )
        return
    if msg.content_type == 'photo':

        file_name = f'{msg.from_user.id}_2.jpg'

        if os.path.exists(f'static/users_photo/{file_name}'):
            os.remove(f'static/users_photo/{file_name}')
        path = f'static/users_photo/{file_name}'
        await msg.photo[-1].download(path)

        async with state.proxy() as data:
                data['screenshot2'] = path
        if await update_screenshots(str(msg.from_user.id), data["screenshot1"], data["screenshot2"]):
            await bot.send_message(
                msg.from_user.id, 
                'Успешно обновлено!',
                )
        else:
            await bot.send_message(
                msg.from_user.id,
                'Что-то пошло не так, уже чиним...'
            )
        await state.finish()
        await my_profile(msg)
    



@dp.message_handler(Text('𝟒'))
async def state_change_description(msg: types.Message):
    await bot.send_message(
        msg.from_user.id,
        'Расскажи о своем игровом опыте и кого хочешь найти. Это поможет лучше подобрать тебе тимейтов. ',
        reply_markup=change_profile_description_cancel()
    )
    await Change_description.description.set()

@dp.message_handler(state=Change_description.description)
async def state_change_description_step2(msg: types.Message, state: FSMContext):
    if msg.text == '❌ Отмена':
        await state.finish()
        await my_profile(msg)
    else:
        if len(msg.text) < 800:
            if ('https://' not in msg.text) and ('.com' not in msg.text) and ('.ru' not in msg.text) and ('t.me' not in msg.text):
                user_id = str(msg.from_user.id)
                async with state.proxy() as data:
                    data["description"] = msg.text
                if await change_description_by_id(user_id, data["description"]):
                    await bot.send_message(
                        msg.from_user.id,
                        'Успешно обновлено!'
                    )
                else:
                    await bot.send_message(
                        msg.from_user.id,
                        'Что-то пошло не так, уже чиним...'
                    )
                await state.finish()
                await my_profile(msg)
            else:
                await bot.send_message(
                    msg.from_user.id,
                    'Этот текст нарушает правила, введите другой'
                )
                return
        else:
            await bot.send_message(
                msg.from_user.id,
                'Это описание слишком длинное, попробуйте еще раз (до 800 символов)'
            )
            return
            
        

@dp.message_handler(Text("3"))
async def disable_active(msg: types.Message):
    user_id = str(msg.from_user.id)
    user = await get_user_by_id(user_id)
    await update_active_to_false(user_id)
    if user_id in dict_of_profiles:
        del dict_of_profiles[user_id]
    for users in dict_of_profiles:
        if user_id in dict_of_profiles[users]["profiles_list"]:
            dict_of_profiles[users]["profiles_list"].remove(user_id)
    await bot.send_message(
        msg.from_user.id,
        f'Рад был помочь,\nНадеюсь, ты нашел кого-то благодаря мне',
        reply_markup=reminder_kb()
    )
    
@dp.message_handler(Text('🚀 Cмотреть анкеты'))
async def enable_active(msg: types.Message):
    await update_active_to_true(str(msg.from_user.id))
    await menu(msg)

@dp.message_handler(Text('4'))
async def call_referal_code(msg: types.Message):
    try:
        link = await get_start_link(str(msg.from_user.id), encode=True)
        users = await get_ref_stat()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'User_id'
        sheet['B1'] = 'Referal'
        datas = []
        for user in users:
            try:
                promocode = decode_payload(user.promocode)
                datas.append((user.user_id, promocode))
            except:
                continue
        for row_index, (user_id, referal) in enumerate(datas, start=2):  # Начинаем с 2 строки, так как 1 строка занята заголовками
            sheet[f'A{row_index}'] = user_id
            sheet[f'B{row_index}'] = referal
        file_name = 'static/utils_data/users_referals.xlsx'
        workbook.save(file_name)
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        count = 0
        for cell in sheet['B']:
            if cell.value == str(msg.from_user.id):
                count += 1


        await bot.send_message(
            msg.from_user.id,
            f'Ты пригласил: <b>{count} пользователей</b>.\n\nТвой реферальный промокод:\n\n{link}\n\nПри регистрации нового пользователя по твоему реферальному промокоду, ваша анкета и анкета пользоватлея, который зарегистрировался получат буст на 24 часа!',
            parse_mode='HTML'
        )
        await menu(msg)
    except:
        await bot.send_message(
            msg.from_user.id,
            'Что-то пошло не так...'
        )
###############

########################################################################################
    
################################### Смотреть анкеты ######################################
    

async def search_love_reg(msg: types.Message):
    user_id = str(msg.from_user.id)
    await update_active_to_true(user_id)
    data = await get_user_by_id(user_id)
    if not data["is_blocked"]:
        if user_id not in dict_of_profiles:
            list_of_profiles = await get_list_of_profiles(
                user_id,
                )
            dict_of_profiles[user_id] = {
                    "profiles_list": list_of_profiles,
                    "last_activity": int(time.time()),
                    "like": [],
                    "history_dislike": [],
                    "who_like": [],
                    "activity": 0
                }
            for user in dict_of_profiles:
                if user != user_id:
                    new_user = await get_user_by_id(user_id)
                    old_user = await get_user_by_id(user)
                    if (new_user != 'User not found') and (old_user != 'User not found'):
                        if len(dict_of_profiles[user]['profiles_list']) != 0:
                            place = random.randint(0, len(dict_of_profiles[user]['profiles_list']))
                            dict_of_profiles[user]['profiles_list'].insert(place, user_id)
                        else:
                            dict_of_profiles[user]['profiles_list'].append(user_id)
                                    
            await search_love_step1(msg)
        else:
            if len(dict_of_profiles[user_id]["profiles_list"]) == 0:
                list_of_profiles = await get_list_of_profiles(
                    user_id,
                    )
                if len(list_of_profiles) != 0:
                    dict_of_profiles[user_id]["profiles_list"] = list_of_profiles
                    await search_love_step1(msg)
                else:
                    await bot.send_message(
                        msg.from_user.id,
                        'На данный момент у нас нет подходящих анкет для вас'
                    )
            else:
                await search_love_step1(msg)
    elif data["is_blocked"]:
        await bot.send_message(
            msg.from_user.id,
            'Ваша анкета была заблокирована'
        )



async def search_love_step1(msg: types.Message):
    try:
        user_id = str(msg.from_user.id)
        list_of_profiles = dict_of_profiles[user_id]["profiles_list"]

        if list_of_profiles == 'Пользователи не найдены':
            await bot.send_message(
                msg.from_user.id,
                'На данный момент у нас нет подходящих анкет для вас',
                reply_markup=search_kb()
            )
            
        else:
            if len(list_of_profiles) != 0:
                next_profile_id = list_of_profiles[-1]
                self_data = await get_user_by_id(str(msg.from_user.id))
                next_profile = await get_user_by_id(next_profile_id)
                if next_profile != 'User not found':
                    media = types.MediaGroup()
                    roles = '\nЛюбимые роли:\n'
                    if next_profile["role1"] != '':
                        roles += f'{next_profile["role1"]}\n'
                    if next_profile["role2"] != '':
                        roles += f'{next_profile["role2"]}\n'
                    if next_profile["role3"] != '':
                        roles += f'{next_profile["role3"]}\n'
                    
                    if (next_profile["screenshot1"] != '') and (next_profile["screenshot2"] != ''):
                        media.attach_photo(types.InputFile(
                            next_profile["screenshot1"]),
                            f'{next_profile["name"]}\n{roles}\n{next_profile["description"]}')
                        media.attach_photo(types.InputFile(next_profile["screenshot2"]))
                        await bot.send_media_group(
                            msg.from_user.id, 
                            media=media,
                        )

                        
                    elif (next_profile["screenshot1"] != '') and (next_profile["screenshot2"] == ''):
                        media.attach_photo(types.InputFile(
                            next_profile["screenshot1"]),
                            f'{next_profile["name"]}\n{roles}\n{next_profile["description"]}')
                        await bot.send_media_group(
                            msg.from_user.id, 
                            media=media,
                        )

                    else:
                        await bot.send_message(
                            msg.from_user.id,
                            f'{next_profile["name"]}\n{roles}\n{next_profile["description"]}'
                        )
                else:
                    await bot.send_message(
                            msg.from_user.id,
                            'На данный момент у нас нет подходящих анкет для вас',
                            reply_markup=search_kb()
                        )
            else:
                await search_love_reg(msg)
    except: 
        if len(dict_of_profiles[user_id]["profiles_list"]) != 0:
            dict_of_profiles[user_id]["profiles_list"].pop()
        await search_love_reg(msg)
    
@dp.message_handler(Text('❤️'))
async def like_main(msg: types.Message):
    user_id = str(msg.from_user.id)
    try:
        user = await get_user_by_id(user_id)
        if not user["is_blocked"]:
            if dict_of_profiles[user_id]["activity"] < 10000:
                dict_of_profiles[user_id]["last_activity"] = int(time.time())
                list_of_profiles = dict_of_profiles[user_id]["profiles_list"]
                like = list_of_profiles[-1]
                if user_id not in dict_of_profiles[list_of_profiles[-1]]["who_like"]:
                    dict_of_profiles[list_of_profiles[-1]]["who_like"].append(user_id)
                    
                    who_len = len(dict_of_profiles[list_of_profiles[-1]]["who_like"])
                    dict_of_profiles[user_id]["profiles_list"].pop()
                    await bot.send_message(
                        like,
                        'Вы понравились 1 человеку, показать его?' if who_len == 1 else f'Вы понравились {who_len} людям, показать их?',
                        reply_markup=show_like_kb()
                    )

                    await bot.send_message(
                        msg.from_user.id,
                        'Сердечко успешно отправлено, ждем ответа пользователя...',
                    )
                elif user_id in dict_of_profiles[list_of_profiles[-1]]["who_like"]:
                    dict_of_profiles[user_id]["profiles_list"].pop()
                    await bot.send_message(
                        msg.from_user.id,
                        'Мы до сих пор ждем ответа на прошлую реакцию...'
                    )

                await search_love_step1(msg)
            else:
                await bot.send_message(
                    msg.from_user.id,
                    'Вы привысили количество реакций, подождите немного для продолжения поиска...'
                )
                await menu(msg)
        elif user["is_blocked"]:
            await bot.send_message(
                msg.from_user.id,
                'Ваша анкета была заблокирована'
            )
    except Exception as e:
        await bot.send_message(
            msg.from_user.id,
            'Что-то пошло не так, уже чиним...'
        )
        logging.error(user_id, exc_info=True)
        await search_love_step1(msg)
    
@dp.message_handler(Text('👎'))
async def dislike_main(msg: types.Message):
    user_id = str(msg.from_user.id)
    user = await get_user_by_id(user_id)
    if not user["is_blocked"]:
        dict_of_profiles[user_id]["last_activity"] = int(time.time())
        if len(dict_of_profiles[user_id]["history_dislike"]) == 5:
            dict_of_profiles[user_id]["history_dislike"].pop(0)
        if dict_of_profiles[user_id]["profiles_list"][-1] not in dict_of_profiles[user_id]["history_dislike"]:
            dict_of_profiles[user_id]["history_dislike"].append(dict_of_profiles[user_id]["profiles_list"][-1])
        dict_of_profiles[user_id]["profiles_list"].pop()
        await search_love_step1(msg)
    elif user["is_blocked"]:
        await bot.send_message(
            msg.from_user.id,
            'Ваша анкета была заблокирована'
        )



@dp.message_handler(Text('🚀 Показать'))
async def show_like(msg: types.Message):
    user_id = str(msg.from_user.id)
    user = await get_user_by_id(user_id)
    if not user["is_blocked"]:
        dict_of_profiles[user_id]["last_activity"] = int(time.time())
        who_like = dict_of_profiles[str(msg.from_user.id)]["who_like"][-1]
        profile = await get_user_by_id(who_like)
        user = await get_user_by_id(user_id)

        if user != 'User not found':
            media = types.MediaGroup()
            await bot.send_message(
                    msg.from_user.id,
                    'Кому-то понравилась ваша анкета:',
                    reply_markup=like_kb()
                )
            roles = '\nЛюбимые роли:\n'
            if profile["role1"] != '':
                roles += f'{profile["role1"]}\n'
            if profile["role2"] != '':
                roles += f'{profile["role2"]}\n'
            if profile["role3"] != '':
                roles += f'{profile["role3"]}\n'

            if (profile["screenshot1"] != '') and (profile["screenshot2"] != ''):
                media.attach_photo(types.InputFile(
                    f'static/users_photo/{who_like}_1.jpg'),
                    f'{profile["name"]}\n{roles}\n{profile["description"]}')
                media.attach_photo(types.InputFile(f'static/users_photo/{who_like}_2.jpg'))
                await bot.send_media_group(
                    msg.from_user.id, 
                    media=media,
                )
                

                
            elif (profile["screenshot1"] != '') and (profile["screenshot2"] == ''):
                media.attach_photo(types.InputFile(
                    f'static/users_photo/{who_like}_1.jpg'),
                    f'{profile["name"]}\n{roles}\n{profile["description"]}')
                await bot.send_media_group(
                    msg.from_user.id, 
                    media=media,
                )
                

            else:
                await bot.send_message(
                    msg.from_user.id,
                    f'{profile["name"]}\n{roles}\n{profile["description"]}'
                )
    elif user["is_blocked"]:
        await bot.send_message(
            msg.from_user.id,
            'Ваша анкета была заблокирована'
        )
    

@dp.message_handler(Text('💜'))
async def like_liked(msg: types.Message):
    global love_activity
    user_id = str(msg.from_user.id)
    user = await get_user_by_id(user_id)
    if not user["is_blocked"]:
        dict_of_profiles[user_id]["last_activity"] = int(time.time())
        who_like = dict_of_profiles[str(msg.from_user.id)]["who_like"]
        user = await get_user_by_id(who_like[-1])
        user_like = await get_user_by_id(str(msg.from_user.id))
        who_like_id = dict_of_profiles[str(msg.from_user.id)]["who_like"][-1]
        dict_of_profiles[str(msg.from_user.id)]["who_like"].pop()
        love_activity[0] += 1
        roles = '\nЛюбимые роли:\n'
        if user_like["role1"] != '':
            roles += f'{user_like["role1"]}\n'
        if user_like["role2"] != '':
                roles += f'{user_like["role2"]}\n'
        if user_like["role3"] != '':
                roles += f'{user_like["role3"]}\n'

        media = types.MediaGroup()

        if (user_like["screenshot1"] != '') and (user_like["screenshot2"] != ''):
                media.attach_photo(types.InputFile(
                    f'static/users_photo/{str(msg.from_user.id)}_1.jpg'),
                    caption=f'Тобой заинтересовались, скорее начните общаться: @{user_like["user_name"]}\n\nИгровой id: <code>{user_like["game_id"]}</code>\n\n{user_like["name"]}\n{roles}\n{user_like["description"]}', parse_mode="HTML")
                media.attach_photo(types.InputFile(f'static/users_photo/{str(msg.from_user.id)}_2.jpg'))
                await bot.send_media_group(
                    who_like_id, 
                    media=media
                )
                await bot.send_message(
                    who_like_id,
                    f'Давай смотреть анкеты дальше\n\n1. Смотреть анкеты.\n2. Моя анкета.\n3. Я больше не хочу никого искать.\n***\n4. Пригласи друзей - получи больше лайков 😎.',
                    reply_markup=menu_kb()
                )
                

                
        elif (user_like["screenshot1"] != '') and (user_like["screenshot2"] == ''):
                media.attach_photo(types.InputFile(
                    f'static/users_photo/{str(msg.from_user.id)}_1.jpg'),
                    caption=f'Тобой заинтересовались, скорее начните общаться: @{user_like["user_name"]}\n\nИгровой id: <code>{user_like["game_id"]}</code>\n\n{user_like["name"]}\n{roles}\n{user_like["description"]}', parse_mode="HTML")
                await bot.send_media_group(
                    who_like_id, 
                    media=media,
                )
                await bot.send_message(
                    who_like_id,
                    f'Давай смотреть анкеты дальше\n\n1. Смотреть анкеты.\n2. Моя анкета.\n3. Я больше не хочу никого искать.\n***\n4. Пригласи друзей - получи больше лайков 😎.',
                    reply_markup=menu_kb()
                )
                

        else:
                await bot.send_message(
                    who_like_id,
                    f'Тобой заинтересовались, скорее начните общаться: @{user_like["user_name"]}\n\nИгровой id: <code>{user_like["game_id"]}</code>\n\n{user_like["name"]}\n{roles}\n{user_like["description"]}', 
                    parse_mode="HTML"
                )
                await bot.send_message(
                    who_like_id,
                    f'Давай смотреть анкеты дальше\n\n1. Смотреть анкеты.\n2. Моя анкета.\n3. Я больше не хочу никого искать.\n***\n4. Пригласи друзей - получи больше лайков 😎.',
                    reply_markup=menu_kb()
                )
        # await bot.send_message(
        #     who_like_id,
        #     f'Тобой заинтересовались, скорее начните общаться: @{user_like["user_name"]}\n\nИгровой id: {user_like["game_id"]}\n\nДавай смотреть анкеты дальше\n\n1. Смотреть анкеты.\n2. Моя анкета.\n3. Я больше не хочу никого искать.\n***\n4. Пригласи друзей - получи больше лайков 😎.',
        #     reply_markup=menu_kb()
        # )
        await bot.send_message(
            msg.from_user.id, 
            f'Скорее начинайте общаться: @{user["user_name"]}\n\nИгровой id: <code>{user["game_id"]}</code>',
            parse_mode='HTML'
        )
        
        if len(dict_of_profiles[str(msg.from_user.id)]["who_like"]) != 0:
            await show_like(msg)
        else:
            await bot.send_message(
                msg.from_user.id,
                'На этом все, давай смотреть анкеты дальше',
            )
            await menu(msg)
    elif user["is_blocked"]:
        await bot.send_message(
            msg.from_user.id,
            'Ваша анкета была заблокирована'
        )

@dp.message_handler(Text('👎🏻'))
async def dislike_liked(msg: types.Message):
    user_id = str(msg.from_user.id)
    user = await get_user_by_id(user_id)
    if not user["is_blocked"]:
        dict_of_profiles[user_id]["last_activity"] = int(time.time())
        dict_of_profiles[str(msg.from_user.id)]["who_like"].pop()
        if len(dict_of_profiles[str(msg.from_user.id)]["who_like"]) != 0:
            await show_like(msg)
        else:
            await bot.send_message(
                msg.from_user.id,
                'На этом все, давай смотреть анкеты дальше',
            )
            await menu(msg)
    elif user["is_blocked"]:
        await bot.send_message(
            msg.from_user.id,
            'Ваша анкета была заблокирована'
        )
@dp.message_handler(Text('❗️'))
@dp.message_handler(Text('⚠️'))
async def search_report(msg: types.Message):
    user_id = str(msg.from_user.id)
    dict_of_profiles[user_id]["last_activity"] = int(time.time())
    cached_data[user_id] = msg.text
    await bot.send_message(
        msg.from_user.id, 
        'Укажите причину жалобы:',
        reply_markup=report_kb()
    )

@dp.callback_query_handler(lambda c: 'report' in c.data and c.data != 'report:other')
async def report_callback(callback_query: types.CallbackQuery):
    user_id = str(callback_query.from_user.id)
    report = callback_query.data.split(':')[1]
    if cached_data[user_id] == '❗️':
        report_user_id = dict_of_profiles[user_id]["who_like"][-1] 
        dict_of_profiles[user_id]["who_like"].pop()
    elif cached_data[user_id] == '⚠️':      
        report_user_id = dict_of_profiles[user_id]["profiles_list"][-1] 
        dict_of_profiles[user_id]["profiles_list"].pop()
    if report != 'cancel':
        list_of_admins = await get_list_of_admins()
        report_data = await get_user_by_id(report_user_id)
        roles = '\nЛюбимые роли:\n'
        if report_data["role1"] != '':
            roles += f'{report_data["role1"]}\n'
        if report_data["role2"] != '':
            roles += f'{report_data["role2"]}\n'
        if report_data["role3"] != '':
            roles += f'{report_data["role3"]}\n'
        for admin_id in list_of_admins:
            media = types.MediaGroup()
            if (report_data["screenshot1"] != '') and (report_data["screenshot2"] != ''):
                media.attach_photo(types.InputFile(
                    f'static/users_photo/{report_data["user_id"]}_1.jpg'),
                    caption=f'Пользователь {user_id} подал жалобу на {report_user_id} по причине "{report}"\n\n{report_data}')
                media.attach_photo(types.InputFile(f'static/users_photo/{report_data["user_id"]}_2.jpg'))
                await bot.send_media_group(
                    admin_id, 
                    media=media,
                )
                # await bot.send_message(
                #     admin_id,
                #     f'👆👆👆👆👆👆👆\nПользователь {user_id} подал жалобу на {report_user_id} по причине "{report}"\n\n{report_data}'
                # ) 

                
            elif (report_data["screenshot1"] != '') and (report_data["screenshot2"] == ''):
                media.attach_photo(types.InputFile(
                    f'static/users_photo/{report_data["user_id"]}_1.jpg'), caption=f'Пользователь {user_id} подал жалобу на {report_user_id} по причине "{report}"\n\n{report_data}')
                await bot.send_media_group(
                    admin_id, 
                    media=media,
                )
                

            else:
                await bot.send_message(
                    admin_id,
                    f'Пользователь {user_id} подал жалобу на {report_user_id} по причине "{report}"\n\n{report_data}'
                )   
            # await bot.send_message(
            #     admin_id,
            #     f'Пользователь {user_id} подал жалобу на {report_user_id} по причине "{report}"\n\n{report_data}'
            # )
        await bot.send_message(
            callback_query.from_user.id, 
            'Жалоба успешно отправлена'
        )     
    await menu(callback_query)

@dp.callback_query_handler(lambda c: c.data == 'report:other')
async def report_callback_other(callback_query: types.CallbackQuery):
    await bot.send_message(
        callback_query.from_user.id,
        'Опишите причину подачи жалобы'
    )
    await ReportUserOther.cause.set()

@dp.message_handler(state=ReportUserOther.cause)
async def report_callback_other_state(msg: types.Message, state: FSMContext):
    user_id = str(msg.from_user.id)
    async with state.proxy() as data:
        data["cause"] = msg.text
    report_user_id = dict_of_profiles[user_id]["profiles_list"][-1] 
    dict_of_profiles[user_id]["profiles_list"].pop()
    list_of_admins = await get_list_of_admins()
    report_data = await get_user_by_id(report_user_id)
    roles = '\nЛюбимые роли:\n'
    if report_data["role1"] != '':
        roles += f'{report_data["role1"]}\n'
    if report_data["role2"] != '':
        roles += f'{report_data["role2"]}\n'
    if report_data["role3"] != '':
        roles += f'{report_data["role3"]}\n'
    for admin_id in list_of_admins:
        media = types.MediaGroup()
        if (report_data["screenshot1"] != '') and (report_data["screenshot2"] != ''):
            media.attach_photo(types.InputFile(
                f'static/users_photo/{report_data["user_id"]}_1.jpg'), caption=f'Пользователь {user_id} подал жалобу на {report_user_id} по причине "{data["cause"]}"\n\n{report_data}')
            media.attach_photo(types.InputFile(f'static/users_photo/{report_data["user_id"]}_2.jpg'))
            await bot.send_media_group(
                admin_id, 
                media=media,
            )
            # await bot.send_message(
            #     admin_id,
            #     f'👆👆👆👆👆👆👆\nПользователь {user_id} подал жалобу на {report_user_id} по причине "{data["cause"]}"\n\n{report_data}'
            # ) 

                
        elif (report_data["screenshot1"] != '') and (report_data["screenshot2"] == ''):
            media.attach_photo(types.InputFile(
                f'static/users_photo/{report_data["user_id"]}_1.jpg'), caption=f'Пользователь {user_id} подал жалобу на {report_user_id} по причине "{data["cause"]}"\n\n{report_data}')
            await bot.send_media_group(
                admin_id, 
                media=media,
            )
            # await bot.send_message(
            #     admin_id,
            #     f'👆👆👆👆👆👆👆\nПользователь {user_id} подал жалобу на {report_user_id} по причине "{data["cause"]}"\n\n{report_data}'
            # ) 

        else:
            await bot.send_message(
                admin_id,
                f'Пользователь {user_id} подал жалобу на {report_user_id} по причине "{data["cause"]}"\n\n{report_data}'
            ) 
    await state.finish()
    await bot.send_message(
        msg.from_user.id, 
        'Жалоба успешно отправлена'
    )  
    await menu(msg)

    

#########################################################################################################################

##################################### Панель администратора ####################################################
    

@dp.message_handler(commands='admin')
async def login_admin(msg: types.Message):
    user_id = str(msg.from_user.id)
    list_of_admins = await get_list_of_admins()
    if user_id in list_of_admins:
        await bot.send_message(
            msg.from_user.id, 
            'Ты попал в админ панель, вот список команд:\n\
"/block_user"\n\
"/get_user_by_id"\n\
"/unblock_user"\n\
"/get_statistic"\n\
"/spam"\n\
"/get_user_by_photo"\n\
"/send_message_from_user_by_id"\n\
"/add_admin"\n\
"/delete_admin"\n\
"/get_backups"\n\
"/create_backups"\n\
"/get_ref_stat"\n\
"/get_ref_by_id"\n\
"/get_love_stat"\n\
"/delete_user"'
        )



@dp.message_handler(commands='get_love_stat')
async def admin_get_love_stat(msg: types.Message):
    user_id = str(msg.from_user.id)
    list_of_admins = await get_list_of_admins()
    if user_id in list_of_admins:
        await bot.send_message(
            msg.from_user.id,
            f'Взаимных симпатий за последние 24 часа: {love_activity[0]}'
        )




@dp.message_handler(commands='get_ref_by_id')
async def admin_get_ref_by_id(msg: types.Message, state: FSMContext):
    user_id = str(msg.from_user.id)
    list_of_admins = await get_list_of_admins()
    if user_id in list_of_admins:
        try:
            await bot.send_message(
                msg.from_user.id,
                'Введите id пользователя, для отмена введите "Отмена"'
                )
            await AdminGetRefStat.user_id.set()
        except: 
            await bot.send_message(
                msg.from_user.id,
                'Что-то пошло не так...'
            )

@dp.message_handler(state=AdminGetRefStat.user_id)
async def admin_get_ref_by_id_state(msg: types.Message, state: FSMContext):
    try:
        if msg.text.lower() == 'отмена':
            async with state.proxy() as data:
                data["user_id"] = msg.text
            await state.finish()
            await bot.send_message(
                msg.from_user.id,
                'Действие отменено!'
            )
        else:
            async with state.proxy() as data:
                data["user_id"] = msg.text
            if await has_register(data["user_id"]):
                users = await get_ref_stat()
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet['A1'] = 'User_id'
                sheet['B1'] = 'Referal'
                datas = []
                for user in users:
                    try:
                        promocode = decode_payload(user.promocode)
                        datas.append((user.user_id, promocode))
                    except:
                        continue
                for row_index, (user_id, referal) in enumerate(datas, start=2):  # Начинаем с 2 строки, так как 1 строка занята заголовками
                    sheet[f'A{row_index}'] = user_id
                    sheet[f'B{row_index}'] = referal
                file_name = 'static/utils_data/users_referals.xlsx'
                workbook.save(file_name)
                workbook = openpyxl.load_workbook(file_name)
                sheet = workbook.active
                count = 0
                for cell in sheet['B']:
                    if cell.value == data['user_id']:
                        count += 1
                await state.finish()
                await bot.send_message(
                    msg.from_user.id,
                    f'Количество пользователей, которое прошли по реферальной ссылке этого пользователя: {count}'
                )
            else:
                await state.finish()
                await bot.send_message(
                    msg.from_user.id,
                    'Похоже такого пользователя не существует'
                )
    except Exception as e:
        print(e)
        await state.finish()
        await bot.send_message(
            msg.from_user.id,
            'Что-то пошло не так...'
        )


@dp.message_handler(commands='create_backups')
async def admin_create_backups(msg: types.Message):
    user_id = str(msg.from_user.id)
    list_of_admins = await get_list_of_admins()
    if user_id in list_of_admins:
        try:
            await dump_dict_of_profiles(dict_of_profiles)
            await backup_bd()
            await bot.send_message(
                msg.from_user.id,
                'Бэкап успешно создан'
            )
        except Exception as e:
            logging.error("Ошибка при бэкапе данных", exc_info=True)
            await bot.send_message(
                msg.from_user.id,
                'что-то пошло не так'
            )



@dp.message_handler(commands='get_backups')
async def admin_get_backups(msg: types.Message):
    user_id = str(msg.from_user.id)
    list_of_admins = await get_list_of_admins()
    if user_id in list_of_admins:
        with open('static/backups/database.sql', 'rb') as file:
            await bot.send_document(msg.from_user.id, document=file)
        with open('static/backups/dump.json', 'rb') as file:
            await bot.send_document(msg.from_user.id, document=file)


@dp.message_handler(commands='add_admin')
async def admin_add_admin(msg: types.Message):
    user_id = str(msg.from_user.id)
    if user_id == main_admin:
        await AdminAddAdmin.user_id.set()
        await bot.send_message(
            msg.from_user.id,
            'Введите id пользователя'
        )

@dp.message_handler(state=AdminAddAdmin.user_id)
async def state_admin_add_admin(msg: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data["user_id"] = msg.text
    await add_admin_db(user_id=data["user_id"])
    await bot.send_message(
        msg.from_user.id,
        'Админ успешно добавлен'
    )
    await state.finish()

@dp.message_handler(commands='delete_user')
async def admin_delete_user(msg: types.Message):
    user_id = str(msg.from_user.id)
    list_of_admins = await get_list_of_admins()
    if user_id in list_of_admins:
        await AdminDeleteUser.user_id.set()
        await bot.send_message(
            msg.from_user.id,
            'Введите id пользователя, для отмены введите "отмена"'
        )            

@dp.message_handler(state=AdminDeleteUser.user_id)
async def state_admin_delete_user(msg: types.Message, state: FSMContext):
    try:
        if msg.text.lower() == 'отмена':
            await state.finish()
            await bot.send_message(
                msg.from_user.id,
                'Действие отменено'
            )
        else:
            async with state.proxy() as data:
                data["user_id"] = msg.text
            if await delete_user(data['user_id']) :
                if data["user_id"] in dict_of_profiles:
                    del dict_of_profiles[data["user_id"]]
                for user in dict_of_profiles:
                    if data["user_id"] in dict_of_profiles[user]["profiles_list"]:
                        dict_of_profiles[user]["profiles_list"].remove(data["user_id"])
                await bot.send_message(
                    int(data['user_id']),
                    'Ваша анкета была удалена, введите /start для активации анкеты'
                )
                await bot.send_message(
                    msg.from_user.id,
                    'Анкета успешно удалена!'
                )
                await state.finish()
            else:
                await bot.send_message(
                    msg.from_user.id,
                    'Что-то пошло не так'
                )
                await state.finish()
    except Exception as e:
        await bot.send_message(
            msg.from_user.id,
            'Что-то пошло не так...'
        )




@dp.message_handler(commands='clear_dict')
async def admin_clear_dict(msg: types.Message):
    try:
        user_id = str(msg.from_user.id)
        list_of_admins = await get_list_of_admins()
        del_list = []
        if user_id in list_of_admins:
            for user in dict_of_profiles:
                user_db = await get_user_by_id(user)
                if (user_db == 'User not found') or (user_db == 'Error'):
                    del_list.append(user)
            print(del_list)
            if len(del_list) != 0:
                for user in dict_of_profiles:
                    for user_id in del_list:
                        if user_id in dict_of_profiles[user]["profiles_list"]:
                            dict_of_profiles[user]["profiles_list"].remove(user_id)

            await bot.send_message(
                msg.from_user.id,
                'Успешно'
            )
    except Exception as e:
        print(e)
        await bot.send_message(
            msg.from_user.id,
            'Что-то пошло не так...'
        )


@dp.message_handler(commands='delete_admin')
async def admin_delete_admin(msg: types.Message):
    user_id = str(msg.from_user.id)
    list_of_admins = await get_list_of_admins()
    if user_id in list_of_admins:
        if user_id == main_admin:
            await AdminDeleteAdmin.user_id.set()
            await bot.send_message(
                msg.from_user.id,
                'Введите id пользователя'
            )

@dp.message_handler(state=AdminDeleteAdmin.user_id)
async def state_admin_delete_admin(msg: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data["user_id"] = msg.text
    await delete_admin_db(user_id=data["user_id"])
    await bot.send_message(
        msg.from_user.id,
        'Админ успешно удален'
    )
    await state.finish()



@dp.message_handler(commands='send_message_from_user_by_id')
async def admin_send_message_from_user_by_id(msg: types.Message):
    user_id = str(msg.from_user.id)
    list_of_admins = await get_list_of_admins()
    if user_id in list_of_admins:
        await AdminSendMessageFromUserById.user_id.set()
        await bot.send_message(
            msg.from_user.id,
            'Введите user_id пользователя, которому хотите отправить сообщение'
        )

@dp.message_handler(state=AdminSendMessageFromUserById.user_id)
async def state_admin_send_message_from_user_by_id_step1(msg: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data["user_id"] = msg.text
    await bot.send_message(
        msg.from_user.id,
        'Введите сообщение, которое хотите отправить пользователю'
    )
    await AdminSendMessageFromUserById.next()

@dp.message_handler(state=AdminSendMessageFromUserById.message)
async def state_admin_send_message_from_user_by_id_step2(msg: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data["message"] = msg.text
    await bot.send_message(
        msg.from_user.id,
        'Подтвердите действие "Да/Нет"'
    )
    await AdminSendMessageFromUserById.next()

@dp.message_handler(state=AdminSendMessageFromUserById.confirmation)
async def state_admin_send_message_from_user_by_id_step3(msg: types.Message, state: FSMContext):
    if msg.text.lower() == 'да':
        try:
            async with state.proxy() as data:
                data["confirmation"] = msg.text
            await bot.send_message(
                int(data["user_id"]),
                data["message"],
                reply_markup=reminder_kb()
            )
            await bot.send_message(
                msg.from_user.id,
                'Сообщение успешно отправлено'
            )
        except Exception as e:
            await state.finish()
            await bot.send_message(
                msg.from_user.id,
                'Что-то пошло не так'
            )   
        await state.finish()
    elif msg.text.lower() == 'нет':
        async with state.proxy() as data:
            data["confirmation"] = msg.text
        await state.finish()
        await bot.send_message(
            msg.from_user.id,
            'Действие отменено'
        )
    else:
        await bot.send_message(
            msg.from_user.id,
            'Нет такого варианта ответа, введите "Да/Нет"'
        )
        return



@dp.message_handler(commands=['block_user'])
async def admin_block_user(msg: types.Message):
    user_id = str(msg.from_user.id)
    list_of_admins = await get_list_of_admins()
    if user_id in list_of_admins:
        await AdminBlockUser.user_id.set()
        await bot.send_message(
            msg.from_user.id, 
            'Введите id пользователя, которого хотите заблокировать.\nЕсли вы хотите отменить действие, введите "Отмена"'
        )

@dp.message_handler(state=AdminBlockUser.user_id)
async def state_admin_block_user_step1(msg: types.Message, state: FSMContext):
    if msg.text.lower() != 'отмена':
        async with state.proxy() as data:
            data["user_id"] = msg.text
        await bot.send_message(
            msg.from_user.id, 
            'Введите причину блокировки, если вы хотите отменить действие, введите "отмена"'
        )
        await AdminBlockUser.next()
    else:
        await state.finish()
        await bot.send_message(
            msg.from_user.id,
            'Действие отменено'
        )
        
@dp.message_handler(state=AdminBlockUser.cause)
async def state_admin_block_user_step2(msg: types.Message, state: FSMContext):
    if msg.text.lower() != 'отмена':
            async with state.proxy() as data:
                data["cause"] = msg.text
            await AdminBlockUser.next()
            await bot.send_message(
                msg.from_user.id,
                'Подтвердите действие, введите "Да/Нет"'
            )
    else:
        await state.finish()
        await bot.send_message(
            msg.from_user.id,
            'Действие отменено'
        )

@dp.message_handler(state=AdminBlockUser.confirmation)
async def state_admin_block_user_step3(msg: types.Message, state: FSMContext):
    if msg.text.lower() == 'да':
        async with state.proxy() as data:
            data["confirmation"] = msg.text
        block = await block_user_db(data["user_id"])
        if block:
            blocked_user = await get_user_by_id(data["user_id"])
            if data["user_id"] in dict_of_profiles:
                del dict_of_profiles[data["user_id"]]
            for user in dict_of_profiles:
                if data["user_id"] in dict_of_profiles[user]["profiles_list"]:
                    dict_of_profiles[user]["profiles_list"].remove(data["user_id"])
            await bot.send_message(
                msg.from_user.id, 
                f'Пользователь {data["user_id"]} успешно заблокирован'
            )
            await state.finish()
            await bot.send_message(
                int(data["user_id"]),
                f'Привет, {blocked_user["name"]}, вы были заблокированы по причине:\n\n"{data["cause"]}"\n\nВы можете обратиться в поддержку и оспорить решение: https://t.me/SuportPoTi_bot'
            )
            
        elif not block:
            await bot.send_message(
                msg.from_user.id, 
                'Произошла ошибка, скорее всего пользователя с таким id не существует'
            )
            await state.finish()
    elif msg.text.lower() == 'нет':
        await bot.send_message(
            msg.from_user.id, 
            'Действие отменено'
        )
        await state.finish()
    else:
        await bot.send_message(
            msg.from_user.id, 
            'Нет такого варианта ответа, введите "Да/Нет"'
        )
        return 
    


@dp.message_handler(commands=['unblock_user'])
async def admin_unblock_user(msg: types.Message):
    user_id = str(msg.from_user.id)
    list_of_admins = await get_list_of_admins()
    if user_id in list_of_admins:
        await AdminUnblockUser.user_id.set()
        await bot.send_message(
            msg.from_user.id, 
            'Введите id пользователя, которого хотите разблокировать.\nЕсли вы хотите отменить действие, введите "Отмена"'
        )

@dp.message_handler(state=AdminUnblockUser.user_id)
async def state_admin_unblock_user_step1(msg: types.Message, state: FSMContext):
    if msg.text.lower() != 'отмена':
        async with state.proxy() as data:
            data["user_id"] = msg.text
        await bot.send_message(
            msg.from_user.id, 
            'Подтвердите действие, введите "Да/Нет"'
        )
        await AdminUnblockUser.next()
    else:
        await state.finish()
        await bot.send_message(
            msg.from_user.id,
            'Действие отменено'
        )

@dp.message_handler(state=AdminUnblockUser.confirmation)
async def state_admin_unblock_user_step2(msg: types.Message, state: FSMContext):
    if msg.text.lower() == 'да':
        async with state.proxy() as data:
            data["confirmation"] = msg.text
        unblock = await unblock_user_db(data["user_id"])
        if unblock:
            await bot.send_message(
                msg.from_user.id,
                f'Пользователь {data["user_id"]} успешно разблокирован'
            )
            unblocked_user = await get_user_by_id(data["user_id"])
            await bot.send_message(
                data["user_id"],
                f'Привет, {unblocked_user["name"]}, ваша анкета раблокированна!\n\nВведите "/menu" для продолжения.'
            )


        elif not unblock:
            await bot.send_message(
                msg.from_user.id, 
                'Произошла ошибка, скорее всего пользователя с таким id не существует'
            )
    elif msg.text.lower() == 'нет':
        await state.finish()
        await bot.send_message(
            msg.from_user.id, 
            'Действие отменено'
        )
    else:
        await bot.send_message(
            msg.from_user.id, 
            'Нет такого варианта ответа, введите "Да/Нет"'
        )
        return 
    await state.finish()

            

@dp.message_handler(commands=['get_user_by_id'])
async def admin_state_get_user_by_id(msg: types.Message):
    user_id = str(msg.from_user.id)
    list_of_admins = await get_list_of_admins()
    if user_id in list_of_admins:
        await bot.send_message(
            msg.from_user.id,
            'Введите id пользователя'
        )
        await AdminGetUserById.user_id.set()

@dp.message_handler(state=AdminGetUserById.user_id)
async def state_admin_get_user_by_id(msg: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data["user_id"] = msg.text
    # user_with_anket = await get_user_by_id(data["user_id"], Anketa=True)
    user_data = await get_user_by_id(data["user_id"])
    if user_data != 'User not found' and user_data != 'Error':
        media = types.MediaGroup()
        roles = '\nЛюбимые роли:\n'
        if user_data["role1"] != '':
            roles += f'{user_data["role1"]}\n'
        if user_data["role2"] != '':
            roles += f'{user_data["role2"]}\n'
        if user_data["role3"] != '':
            roles += f'{user_data["role3"]}\n'

        if (user_data["screenshot1"] != '') and (user_data["screenshot2"] != ''):
            media.attach_photo(types.InputFile(
                f'static/users_photo/{data["user_id"]}_1.jpg'),
                f'{user_data["name"]}\n{roles}\n{user_data["description"]}')
            media.attach_photo(types.InputFile(f'static/users_photo/{data["user_id"]}_2.jpg'))
            await bot.send_media_group(
                msg.from_user.id, 
                media=media,
            )

            
        elif (user_data["screenshot1"] != '') and (user_data["screenshot2"] == ''):
            media.attach_photo(types.InputFile(
                f'static/users_photo/{data["user_id"]}_1.jpg'),
                f'{user_data["name"]}\n{roles}\n{user_data["description"]}')
            await bot.send_media_group(
                msg.from_user.id, 
                media=media,
            )

        else:
            await bot.send_message(
                msg.from_user.id,
                f'{user_data["name"]}\n{roles}\n{user_data["description"]}'
            )    
        await bot.send_message(
            msg.from_user.id,
            user_data,
        )
    else:
        await bot.send_message(
            msg.from_user.id,
            'Пользователь не найден'
        )
    await state.finish()

@dp.message_handler(commands=['get_statistic'])
async def admin_get_user_by_admin(msg: types.Message):
    user_id = str(msg.from_user.id)
    list_of_admins = await get_list_of_admins()
    if user_id in list_of_admins:
        users, count = await get_statistic_user_db()
        await bot.send_message(
            msg.from_user.id,
            f'Всего пользователей: {users}\nАктивных пользователей: {count}'
        )


@dp.message_handler(commands=['spam'])
async def admin_spam(msg: types.Message):
    user_id = str(msg.from_user.id)
    list_of_admins = await get_list_of_admins()
    if user_id in list_of_admins:
        await AdminSpamHasPhoto.has_photo.set()
        await bot.send_message(
            msg.from_user.id,
            'Рассылка будет с фото?("Да/Нет")'
        )

@dp.message_handler(state=AdminSpamHasPhoto.has_photo)
async def state_admin_spam_step1(msg: types.Message, state: FSMContext):
    if msg.text.lower() == 'да' or msg.text.lower() == 'нет':
        async with state.proxy() as data:
            data["has_photo"] = msg.text
        await state.finish()
    else:
        await bot.send_message(
            msg.from_user.id,
            'Нет такого варианта ответа'
        )
        return
    if data["has_photo"].lower() == 'да':
        await bot.send_message(
            msg.from_user.id,
            'Отправьте фотографию'
        )
        await AdminSpamWithPhoto.path.set()
    elif data["has_photo"].lower() == 'нет':
        await bot.send_message(
            msg.from_user.id,
            'Введите текст рассылки'
        )
        await AdminSpamOnlyText.spam_text.set()


######## Спам с фото ########
@dp.message_handler(content_types=['photo'], state=AdminSpamWithPhoto)
async def state_admin_spam_with_photo_step1(msg: types.Message, state: FSMContext):
    try:
        file_name = 'spam.jpg'
        path = f'static/spam_photo/{file_name}'
        async with state.proxy() as data:
                data['path'] = path
        await msg.photo[-1].download(path)
        await bot.send_message(
            msg.from_user.id, 
            'Фотография успешно добавлена, введите текст рассылки'
            )
        await AdminSpamWithPhoto.next()
    except Exception as e:
        await bot.send_message(
            msg.from_user.id, 
            'Произошла ошибка, попробуйте отправить фотографию еще раз...'
        )
        return

@dp.message_handler(state=AdminSpamWithPhoto.spam_text)
async def state_admin_spam_with_photo_step2(msg: types.Message, state: FSMContext):
    try:
        async with state.proxy() as data:
            data["spam_text"] = msg.text
        await bot.send_message(
            msg.from_user.id, 
            'Рассылка будет выглядеть вот так:'
        )
        await bot.send_photo(
            msg.from_user.id,
            open(data["path"], 'rb'),
            data["spam_text"],
            parse_mode='HTML',
            
        )
        await bot.send_message(
            msg.from_user.id,
            'Подтвердите действие для начала рассылки "Да/Нет"'
        )
        
        await AdminSpamWithPhoto.next()
    except:
        await bot.send_message(
            msg.from_user.id,
            'Вы допустили ошибку!'
        )
        await state.finish()

@dp.message_handler(state=AdminSpamWithPhoto.confirmation)
async def state_admin_spam_with_photo_step2(msg: types.Message, state: FSMContext):
    if msg.text.lower() == 'да':
        async with state.proxy() as data:
            data["confirmation"] = msg.text
        list_of_users = await get_list_of_users_for_spam_db()
        count = 0
        for user in list_of_users:
            try:
                await bot.send_photo(
                    user,
                    open(data["path"], 'rb'),
                    data["spam_text"],
                    reply_markup=reminder_kb(),
                    parse_mode='HTML'
                )
                count+=1
            except:
                pass
        await state.finish()
        await bot.send_message(
            msg.from_user.id, 
            f'Рассылка успешно отправлена {count} пользователям!'
        )
    elif msg.text.lower() == 'нет':
        await state.finish()
        await bot.send_message(
            msg.from_user.id,
            'Рассылка отменена'
        )
    else:
        await bot.send_message(
            msg.from_user.id,
            'Нет такого варианта ответа, введите "Да/Нет"'
        )
        return

########################

######## Спам без фото ########

@dp.message_handler(state=AdminSpamOnlyText.spam_text)
async def state_admin_spam_only_text_step1(msg: types.Message, state: FSMContext):
    try:
        async with state.proxy() as data:
            data["spam_text"] = msg.text
        await bot.send_message(
            msg.from_user.id,
            'Рассылка будет выглядеть вот так:'
        )
        await bot.send_message(
            msg.from_user.id,
            data["spam_text"],
            parse_mode='HTML',
            disable_web_page_preview=True
        )
        await bot.send_message(
            msg.from_user.id,
            'Подтвердите действие для начала рассылки "Да/Нет"'
        )
        await AdminSpamOnlyText.next()
    except:
        await bot.send_message(
            msg.from_user.id,
            'Вы допустили ошибку!'
        )
        await state.finish()

@dp.message_handler(state=AdminSpamOnlyText.confirmation)
async def state_admin_spam_only_text_step2(msg: types.Message, state: FSMContext):
    if msg.text.lower() == 'да':
        async with state.proxy() as data:
            data["confirmation"] = msg.text
        list_of_users = await get_list_of_users_for_spam_db()
        count = 0
        for user in list_of_users:
            try:
                await bot.send_message(
                    user,
                    data["spam_text"],
                    reply_markup=reminder_kb(),
                    parse_mode='HTML',
                    disable_web_page_preview=True
                )
                count+=1
            except:
                pass
        await state.finish()
        await bot.send_message(
            msg.from_user.id, 
            f'Рассылка успешно отправлена {count} пользователям!'
        )
    elif msg.text.lower() == 'нет':
        await state.finish()
        await bot.send_message(
            msg.from_user.id,
            'Рассылка отменена'
        )
    else:
        await bot.send_message(
            msg.from_user.id,
            'Нет такого варианта ответа, введите "Да/Нет"'
        )
        return

########################

@dp.message_handler(commands=['get_ref_stat'])
async def admin_get_ref_stat(msg: types.Message):
    user_id = str(msg.from_user.id)
    list_of_admins = await get_list_of_admins()
    if user_id in list_of_admins:
        users = await get_ref_stat()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'User_id'
        sheet['B1'] = 'Referal'
        data = []
        for user in users:
            try:
                promocode = decode_payload(user.promocode)
                data.append((user.user_id, promocode))
            except:
                continue
        for row_index, (user_id, referal) in enumerate(data, start=2):  # Начинаем с 2 строки, так как 1 строка занята заголовками
            sheet[f'A{row_index}'] = user_id
            sheet[f'B{row_index}'] = referal
        file_name = 'static/utils_data/users_referals.xlsx'
        workbook.save(file_name)
        with open('static/utils_data/users_referals.xlsx', 'rb') as file:
            await bot.send_document(msg.from_user.id, document=file)

######## Поиск по фото ########
@dp.message_handler(commands=['get_user_by_photo'])
async def admin_get_user_by_photo(msg: types.Message):
    user_id = str(msg.from_user.id)
    list_of_admins = await get_list_of_admins()
    if user_id in list_of_admins:
        await bot.send_message(
            msg.from_user.id,
            'Отправьте фотографию'
        )
        await AdminGetUserByPhoto.path.set()
    
@dp.message_handler(content_types=['photo'], state=AdminGetUserByPhoto.path)
async def state_admin_get_user_by_photo(msg: types.Message, state: FSMContext):
    try:
        file_name = 'search.jpg'
        path = f'static/utils_data/{file_name}'
        async with state.proxy() as data:
                data['path'] = path
        await msg.photo[-1].download(path)
        rezult, user = await compare_images(path)
        if user.isdigit():
            await bot.send_message(
                msg.from_user.id,
                rezult
            )
            anketa = await get_user_by_id(user)
            await bot.send_message(
                msg.from_user.id,
                anketa
            )    
        else:
            await bot.send_message(
                msg.from_user.id,
                rezult
            )
        await state.finish()
    except Exception as e:
        await state.finish()
        await bot.send_message(
            msg.from_user.id, 
            'Похоже такого пользователя не существует в базе данных...'
        )
        await login_admin(msg)


def load_data():
    global dict_of_profiles
    try:
        with open('static/backups/dump.json', 'r') as file:
            dict_of_profiles = json.loads(file.read())
        logging.info('Успешная запись данных из dump в dict_of_profiles')
    except:
        logging.error('Ошибка записи данных из dump в dict_of_profiles')





if __name__ == '__main__':
    start_db()
    load_data()
    scheduler = AsyncIOScheduler()
    logging.getLogger().setLevel(logging.INFO)

    # Настройка обработчика для info-сообщений
    info_handler = logging.FileHandler('utils/main_info.log')
    info_handler.setLevel(logging.INFO)
    info_handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))

    # Настройка обработчика для error-сообщений
    error_handler = logging.FileHandler('utils/main_errors.log')
    error_handler.setLevel(logging.ERROR)
    error_handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))

    # Добавление обработчиков к глобальному логгеру
    logging.getLogger().addHandler(info_handler)
    logging.getLogger().addHandler(error_handler)

    # Запуск потока для бэкапов
    thread_backup_dict_of_profiles = threading.Thread(target=start_schedule, daemon=True, args=(scheduler, dict_of_profiles, bot, love_activity))
    thread_backup_dict_of_profiles.start()
    scheduler.start()
    # Запуск бота
    executor.start_polling(dp, skip_updates=True)






    

    





